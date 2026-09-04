#!/usr/bin/env python3
"""Google口コミ・プレゼント台帳の Excel テンプレートを生成する。

    python tools/build_review_ledger.py 出力先.xlsx [登録済みデータ.json]

第2引数に JSON（口コミの配列）を渡すと、その内容を台帳シートに書き込む。
患者情報を含むデータはリポジトリに置かないこと。
"""
import json
import re
import sys

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

FONT = "Meiryo"
LAST_ROW = 60           # 入力欄をあらかじめ用意しておく行数
SUMMARY_MONTHS = 12

INK = "1F3033"
ACCENT = "0E6E62"
HEAD_FILL = PatternFill("solid", fgColor="0E6E62")
PENDING_FILL = PatternFill("solid", fgColor="FCEAE0")
DONE_FILL = PatternFill("solid", fgColor="E4F1E9")
RULE = Side(style="thin", color="DCE4E1")

COLUMNS = [
    ("対象月", 9),
    ("患者番号", 9),
    ("患者名", 14),
    ("区分", 10),
    ("プレゼント", 18),
    ("状態", 9),
    ("渡した日", 11),
    ("担当", 9),
    ("評価", 6),
    ("投稿日", 11),
    ("Google表示名", 14),
    ("備考", 20),
    ("口コミ内容", 100),
]
CENTERED = (1, 2, 4, 6, 7, 8, 9, 10)   # 中央ぞろえにする列
CAT_COL, GIFT_COL, STATUS_COL = 4, 5, 6

GIFT_ORTHO = "ホワイトニング"
GIFT_GENERAL = "物品1,000円OFF"


def style_header(ws, row=1):
    for col, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 24


def build_ledger(ws, entries):
    # 患者名まで（A〜C列）と見出し行は、右にスクロールしても残す
    ws.freeze_panes = "D2"
    style_header(ws)

    order = ["month", "patientNo", "name", "category", "gift", "status",
             "givenDate", "staff", "stars", "postedAt", "reviewer", "note", "content"]
    for i, e in enumerate(entries):
        r = i + 2
        for c, key in enumerate(order, start=1):
            # プレゼントは区分から自動判定。「その他」で内容がある場合だけ上書きする
            if c == GIFT_COL and not e.get("gift"):
                continue
            value = e.get(key, "")
            if key == "content":
                # 改行を含むと1行表示が崩れるため、1つの段落にまとめる
                value = re.sub(r"\s*\n+\s*", " ", str(value)).strip()
            ws.cell(row=r, column=c, value=value)
        if e.get("comment"):
            ws.cell(row=r, column=2).comment = Comment(e["comment"], "台帳")

    for r in range(2, LAST_ROW + 1):
        if ws.cell(row=r, column=GIFT_COL).value is None:
            ws.cell(row=r, column=GIFT_COL).value = (
                f'=IF(${get_column_letter(CAT_COL)}{r}="矯正中","{GIFT_ORTHO}",'
                f'IF(${get_column_letter(CAT_COL)}{r}="矯正以外","{GIFT_GENERAL}",""))'
            )
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.border = Border(bottom=RULE)
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if c in CENTERED else "left",
            )
        ws.cell(row=r, column=7).number_format = "yyyy/mm/dd"
        ws.cell(row=r, column=10).number_format = "yyyy/mm/dd"
        ws.row_dimensions[r].height = 22   # 1行1件。行の高さをそろえて一覧性を優先する

    ws.auto_filter.ref = f"A1:M{LAST_ROW}"

    dv_cat = DataValidation(type="list", formula1='"矯正中,矯正以外,その他"', allow_blank=True)
    dv_cat.error = "矯正中／矯正以外／その他 から選んでください。"
    dv_status = DataValidation(type="list", formula1='"未渡し,渡し済"', allow_blank=True)
    dv_stars = DataValidation(type="list", formula1='"5,4,3,2,1"', allow_blank=True)
    for dv, col in ((dv_cat, CAT_COL), (dv_status, STATUS_COL), (dv_stars, 9)):
        ws.add_data_validation(dv)
        letter = get_column_letter(col)
        dv.add(f"{letter}2:{letter}{LAST_ROW}")

    # 状態に応じて行全体に色を付ける（未渡しが一目で分かるように）
    body = f"A2:M{LAST_ROW}"
    st = f"${get_column_letter(STATUS_COL)}2"
    ws.conditional_formatting.add(
        body, FormulaRule(formula=[f'{st}="未渡し"'], fill=PENDING_FILL, stopIfTrue=False))
    ws.conditional_formatting.add(
        body, FormulaRule(formula=[f'{st}="渡し済"'], fill=DONE_FILL, stopIfTrue=False))

    # 印刷（ミーティング用に配る場合）: 横向き・幅を1ページに収め、見出し行を各ページに出す
    # 口コミ本文（M列）は紙に載せると読めないので、印刷はL列までにする
    ws.print_area = f"A1:L{LAST_ROW}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:1"


def build_summary(ws, start_month):
    heads = ["対象月", "口コミ件数", "未渡し", "渡し済", "ホワイトニング", "物品1,000円OFF"]
    ws["A1"] = "月別の集計（台帳シートに入力すると自動で更新されます）"
    ws["A1"].font = Font(name=FONT, bold=True, size=12, color=ACCENT)
    for col, title in enumerate(heads, start=1):
        cell = ws.cell(row=3, column=col, value=title)
        cell.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 16 if col == 1 else 14

    year, month = (int(x) for x in start_month.split("-"))
    for i in range(SUMMARY_MONTHS):
        r = 4 + i
        y, m = year + (month - 1 + i) // 12, (month - 1 + i) % 12 + 1
        ws.cell(row=r, column=1, value=f"{y}-{m:02d}")
        ws.cell(row=r, column=2, value=f'=COUNTIF(口コミ台帳!$A$2:$A${LAST_ROW},$A{r})')
        ws.cell(row=r, column=3, value=(
            f'=COUNTIFS(口コミ台帳!$A$2:$A${LAST_ROW},$A{r},'
            f'口コミ台帳!$F$2:$F${LAST_ROW},"未渡し")'))
        ws.cell(row=r, column=4, value=(
            f'=COUNTIFS(口コミ台帳!$A$2:$A${LAST_ROW},$A{r},'
            f'口コミ台帳!$F$2:$F${LAST_ROW},"渡し済")'))
        ws.cell(row=r, column=5, value=(
            f'=COUNTIFS(口コミ台帳!$A$2:$A${LAST_ROW},$A{r},'
            f'口コミ台帳!$D$2:$D${LAST_ROW},"矯正中")'))
        ws.cell(row=r, column=6, value=(
            f'=COUNTIFS(口コミ台帳!$A$2:$A${LAST_ROW},$A{r},'
            f'口コミ台帳!$D$2:$D${LAST_ROW},"矯正以外")'))
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.border = Border(bottom=RULE)
            cell.alignment = Alignment(horizontal="center")


def build_guide(ws):
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 78
    rows = [
        ("使い方", ""),
        ("", ""),
        ("入力する場所", "「口コミ台帳」シートの2行目から、1件につき1行で入力します。行が足りなくなったら下に足してください。"),
        ("プレゼント欄", "区分を選ぶと自動で入ります（矯正中→ホワイトニング／矯正以外→物品1,000円OFF）。"),
        ("", "区分が「その他」のときは、プレゼント欄に直接入力してください（自動の数式は消えます）。"),
        ("渡したかどうか", "状態の欄で「未渡し」「渡し済」を選びます。未渡しはオレンジ、渡し済は緑になります。"),
        ("", "渡し済にしたら、渡した日と担当も入れておくと後から確認できます。"),
        ("月例ミーティング", "見出し行の▼から対象月で絞り込み、状態で「未渡し」だけを表示すると、その場で確認できます。"),
        ("集計", "「月別集計」シートに、月ごとの件数・未渡し・渡し済が自動で出ます。"),
        ("口コミ本文の読み方", "本文は右端のM列にあります。1行に収めているので、全文はセルをクリックして上の入力バーで読めます。"),
        ("印刷", "印刷すると口コミ本文以外のL列までが、横向き1ページ幅で出ます（見出し行は各ページに付きます）。"),
        ("", "全文を表で表示したいときは、M列を選んで［表示形式］→［折り返し］をオンにしてください。"),
        ("", ""),
        ("プレゼントの基準", "矯正中の方 → ホワイトニング"),
        ("", "矯正以外の方 → 物品1,000円OFF"),
        ("", "※基準を変える場合は、この行と「口コミ台帳」シートのプレゼント欄の数式を直してください。"),
        ("", ""),
        ("記入例", "対象月 2026-09 ／ 患者番号 10428 ／ 患者名 佐藤 美咲 ／ 区分 矯正中 ／ "
                  "プレゼント ホワイトニング（自動）／ 状態 渡し済 ／ 渡した日 2026/09/12 ／ 担当 中村"),
    ]
    for i, (label, text) in enumerate(rows, start=1):
        a = ws.cell(row=i, column=1, value=label)
        b = ws.cell(row=i, column=2, value=text)
        a.font = Font(name=FONT, bold=True, size=11 if i > 1 else 14,
                      color=ACCENT if i == 1 else INK)
        b.font = Font(name=FONT, size=10, color=INK)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        if len(text) > 60:
            ws.row_dimensions[i].height = 30


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "口コミプレゼント台帳.xlsx"
    entries = []
    if len(sys.argv) > 2:
        with open(sys.argv[2], encoding="utf-8") as fh:
            entries = json.load(fh)

    wb = Workbook()
    ledger = wb.active
    ledger.title = "口コミ台帳"
    build_ledger(ledger, entries)
    build_summary(wb.create_sheet("月別集計"), entries[0]["month"] if entries else "2026-09")
    build_guide(wb.create_sheet("使い方"))
    wb.save(out)
    print(f"saved: {out} ({len(entries)} 件)")


if __name__ == "__main__":
    main()
