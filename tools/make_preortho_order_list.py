# -*- coding: utf-8 -*-
"""プレオルソ発注リスト（Excel）を生成するスクリプト。

使い方:
    python tools/make_preortho_order_list.py [出力先.xlsx]

既定の出力先は templates/preortho_order_list.xlsx。
選択肢（サイズ・タイプ・色・硬さ）は下部の MASTER を書き換えれば変更できる。
"""

import datetime
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Meiryo"

# --- 配色 ---------------------------------------------------------------
TEAL_DARK = "115E69"   # タイトル帯
TEAL = "17808F"        # 見出し行
TEAL_PALE = "D7EBEF"   # 患者情報グループ
PINK_PALE = "FBDCE6"   # 装置仕様グループ
PINK_TEXT = "8E3B5B"
CREAM_PALE = "FCEBCF"  # 備考グループ
CREAM_TEXT = "8A6116"
BAND = "F3F9FA"        # 縞模様（偶数行）
NOTE_FILL = "FFF9EC"   # 使い方ボックス
NOTE_TEXT = "6B5B3E"
ALERT_FILL = "FFC7CE"  # ロング＋ソフトの警告

# 発注時に選ぶ項目。院内の取り扱いに合わせて自由に増減してよい。
# 硬さは「ハード」を先頭に置くこと（サイズが「ロング」のときはハードのみ選べる仕組みのため）。
MASTER = {
    "サイズ": ["SS", "S", "ロング"],
    "タイプ": [1, 2, 3],
    "色": ["ピンク", "ブルー", "イエロー"],
    "硬さ": ["ハード", "ソフト"],
}
LONG_SIZE = "ロング"  # この値を選んだ行は硬さがハード固定になる

HEADERS = ["日付", "氏名", "サイズ", "タイプ", "色", "硬さ", "備考"]
WIDTHS = [14, 20, 10, 8, 12, 10, 32]
# 見出しの上に置くグループ帯：(表示名, 開始列, 終了列, 背景色, 文字色)
GROUPS = [
    ("患者情報", 1, 2, TEAL_PALE, TEAL_DARK),
    ("装置仕様", 3, 6, PINK_PALE, PINK_TEXT),
    ("その他", 7, 7, CREAM_PALE, CREAM_TEXT),
]
NOTES = [
    "入力は10行目から。色の付いたセルに1件1行で入力します。",
    "サイズ・タイプ・色・硬さはセルを選ぶとプルダウンから選べます。",
    f"サイズが「{LONG_SIZE}」の行は硬さが「ハード」のみ。選択肢は「マスタ」シートで変更できます。",
]
ROWS = 200  # 入力できる行数

HAIR = Side(style="hair", color="BFD4D8")
THIN = Side(style="thin", color="9FC3CA")
MEDIUM = Side(style="medium", color=TEAL)


def cell_border(top=HAIR, bottom=HAIR, left=HAIR, right=HAIR):
    return Border(top=top, bottom=bottom, left=left, right=right)


def build_master(wb):
    ws = wb.create_sheet("マスタ")
    ws["A1"] = "選択肢マスタ"
    ws["A1"].font = Font(name=FONT, size=13, bold=True, color=TEAL_DARK)
    ws["A2"] = "この表を編集すると「発注リスト」のプルダウンに反映されます。"
    ws["A2"].font = Font(name=FONT, size=10, color="595959")

    top = 4
    for col, (name, values) in enumerate(MASTER.items(), start=1):
        letter = get_column_letter(col)
        head = ws.cell(row=top, column=col, value=name)
        head.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        head.fill = PatternFill("solid", fgColor=TEAL)
        head.alignment = Alignment(horizontal="center", vertical="center")
        head.border = cell_border(bottom=MEDIUM)
        for i, value in enumerate(values, start=top + 1):
            c = ws.cell(row=i, column=col, value=value)
            c.font = Font(name=FONT, size=11)
            c.alignment = Alignment(horizontal="center")
            c.border = cell_border(left=THIN, right=THIN, bottom=THIN)
            if (i - top) % 2 == 0:
                c.fill = PatternFill("solid", fgColor=BAND)
        ws.column_dimensions[letter].width = 16
    ws.row_dimensions[top].height = 20

    note_row = top + max(len(v) for v in MASTER.values()) + 2
    note = ws.cell(
        row=note_row,
        column=1,
        value=f"※ サイズが「{LONG_SIZE}」の場合、硬さは「{MASTER['硬さ'][0]}」のみ選べます。"
        "硬さの列は「ハード」を必ず先頭にしてください。",
    )
    note.font = Font(name=FONT, size=10, color=NOTE_TEXT)
    note.fill = PatternFill("solid", fgColor=NOTE_FILL)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    ws.sheet_view.showGridLines = False
    return ws


def build_sheet(wb):
    ws = wb.create_sheet("発注リスト", 0)
    ncols = len(HEADERS)
    last_letter = get_column_letter(ncols)

    for col, width in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # --- タイトル帯（1〜2行目）---
    ws.merge_cells(f"A1:{last_letter}1")
    ws.merge_cells(f"A2:{last_letter}2")
    title = ws["A1"]
    title.value = "プレオルソ 発注リスト"
    title.font = Font(name=FONT, size=16, bold=True, color="FFFFFF")
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    subtitle = ws["A2"]
    subtitle.value = "博多ステラ歯科　業者発注用"
    subtitle.font = Font(name=FONT, size=10, color="CFE7EB")
    subtitle.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for row in (1, 2):
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=TEAL_DARK)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 7  # 余白

    # --- 使い方ボックス（4〜6行目、1行1項目で短く）---
    note_first = 4
    for i, text in enumerate(NOTES):
        row = note_first + i
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1, value=f"▶  {text}")
        cell.font = Font(name=FONT, size=10, color=NOTE_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for col in range(1, ncols + 1):
            target = ws.cell(row=row, column=col)
            target.fill = PatternFill("solid", fgColor=NOTE_FILL)
            target.border = Border(
                left=Side(style="thick", color="E8B84B") if col == 1 else None,
                top=Side(style="hair", color="E7D5AE") if i == 0 else None,
                bottom=Side(style="hair", color="E7D5AE") if i == len(NOTES) - 1 else None,
            )
        ws.row_dimensions[row].height = 19
    ws.row_dimensions[note_first + len(NOTES)].height = 9  # 余白

    # --- グループ帯 ---
    group_row = note_first + len(NOTES) + 1
    for name, start, end, fill, color in GROUPS:
        if end > start:
            ws.merge_cells(start_row=group_row, start_column=start, end_row=group_row, end_column=end)
        cell = ws.cell(row=group_row, column=start, value=name)
        cell.font = Font(name=FONT, size=10, bold=True, color=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(start, end + 1):
            target = ws.cell(row=group_row, column=col)
            target.fill = PatternFill("solid", fgColor=fill)
            target.border = Border(
                left=Side(style="thin", color="FFFFFF") if col == start else None,
                right=Side(style="thin", color="FFFFFF") if col == end else None,
            )
    ws.row_dimensions[group_row].height = 18

    # --- 見出し行 ---
    header_row = group_row + 1
    for col, title_text in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title_text)
        cell.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            bottom=MEDIUM,
            left=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF"),
        )
    ws.row_dimensions[header_row].height = 24

    # --- 入力欄 ---
    first, last = header_row + 1, header_row + ROWS
    centered = (1, 3, 4, 5, 6)
    for row in range(first, last + 1):
        banded = (row - first) % 2 == 1
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name=FONT, size=11)
            cell.alignment = Alignment(
                horizontal="center" if col in centered else "left",
                vertical="center",
                indent=0 if col in centered else 1,
            )
            group_edge = col in (3, 7)
            cell.border = Border(
                top=HAIR,
                bottom=HAIR,
                left=THIN if group_edge else HAIR,
                right=THIN if col == ncols else HAIR,
            )
            if banded:
                cell.fill = PatternFill("solid", fgColor=BAND)
        ws.cell(row=row, column=1).number_format = "yyyy/mm/dd"
        ws.row_dimensions[row].height = 19

    # --- プルダウン（マスタシートの列を参照）---
    columns = {"サイズ": "C", "タイプ": "D", "色": "E", "硬さ": "F"}
    master_first = 5  # マスタシートの1件目の行
    for idx, (name, values) in enumerate(MASTER.items(), start=1):
        letter = get_column_letter(idx)
        if name == "硬さ":
            # サイズが「ロング」の行は先頭1件（ハード）だけを選択肢にする
            source = (
                f"=OFFSET(マスタ!${letter}${master_first},0,0,"
                f'IF($C{first}="{LONG_SIZE}",1,{len(values)}),1)'
            )
        else:
            source = f"=マスタ!${letter}${master_first}:${letter}${master_first + len(values) - 1}"
        dv = DataValidation(type="list", formula1=source, allow_blank=True, showDropDown=False)
        dv.errorTitle = "入力できない値です"
        dv.error = (
            f"サイズが「{LONG_SIZE}」の場合、硬さは「{values[0]}」のみです。"
            if name == "硬さ"
            else f"{name}は「マスタ」シートの選択肢から選んでください。"
        )
        ws.add_data_validation(dv)
        dv.add(f"{columns[name]}{first}:{columns[name]}{last}")

    # --- 念のための保険：ロング＋ソフトの組み合わせが残っていたら赤く塗る ---
    ws.conditional_formatting.add(
        f"F{first}:F{last}",
        FormulaRule(
            formula=[f'AND($C{first}="{LONG_SIZE}",$F{first}<>"",$F{first}<>"{MASTER["硬さ"][0]}")'],
            fill=PatternFill("solid", fgColor=ALERT_FILL),
            font=Font(name=FONT, size=11, bold=True, color="9C0006"),
            stopIfTrue=False,
        ),
    )

    # --- 記入例（実際の発注を入れるときは上書きする）---
    example = [
        datetime.date(2026, 4, 1),
        "博多 太郎",
        "SS",
        2,
        "ピンク",
        "ソフト",
        "記入例：この行は上書きしてください",
    ]
    for col, value in enumerate(example, start=1):
        cell = ws.cell(row=first, column=col, value=value)
        cell.font = Font(name=FONT, size=11, italic=True, color="9AA5A8")
    ws.cell(row=first, column=1).number_format = "yyyy/mm/dd"

    # 表の右側の未使用列は非表示にして、白い余白が出ないようにする
    ws.column_dimensions.group(
        get_column_letter(ncols + 1), "XFD", outline_level=0, hidden=True
    )

    ws.freeze_panes = f"A{first}"
    ws.auto_filter.ref = f"A{header_row}:{last_letter}{last}"
    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f"{header_row}:{header_row}"
    return ws


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "templates/preortho_order_list.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    build_master(wb)
    build_sheet(wb)
    wb.save(out)
    print(f"saved: {out}")


if __name__ == "__main__":
    main()
