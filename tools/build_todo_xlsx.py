#!/usr/bin/env python3
"""data/todo.csv から、月ごとのシートに分けたエクセルを作る。

Chatworkからの取り込みのたびに実行され、data/reception_todo.xlsx を更新する。
"""

import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))
import todo_store  # noqa: E402

OUT = ROOT / "data" / "reception_todo.xlsx"

F = "游ゴシック"
INK = "FF1A2422"; MUTED = "FF5E706C"; FAINT = "FF95A5A1"
TEAL = "FF0F5F55"; SOFT = "FFDCEDE8"; LINE = "FFD9E2DE"; BAND = "FFF7FAF8"
RED = "FFC0392B"; REDBG = "FFFCE9E6"; AMBERBG = "FFFBF0DC"

HEADERS = ["種別", "やること", "期限", "目安時間", "メモ", "✓", "登録日"]
WIDTHS = [13, 56, 15, 12, 40, 7, 15]
HDR = 3
MIN_ROWS = 20

thin = Side(style="thin", color=LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)


def month_key(row):
    src = row.get("created_at") or ""
    try:
        d = date.fromisoformat(src[:10])
    except ValueError:
        d = date.today()
    return f"{d.year}年{d.month}月"


def build_sheet(ws, title, rows):
    ws.merge_cells("A1:G1")
    ws["A1"] = f"受付　TO DO リスト　{title}"
    ws["A1"].font = Font(name=F, size=16, bold=True, color=INK)
    ws.row_dimensions[1].height = 26

    for i, head in enumerate(HEADERS, start=1):
        c = ws.cell(row=HDR, column=i, value=head)
        c.font = Font(name=F, size=10.5, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = box
    ws.row_dimensions[HDR].height = 24

    first = HDR + 1
    last = first + max(len(rows), MIN_ROWS) - 1

    for offset, row in enumerate(rows):
        r = first + offset
        ws.cell(row=r, column=1, value=row.get("kind", ""))
        ws.cell(row=r, column=2, value=row.get("task", ""))
        due = row.get("due") or ""
        if due:
            try:
                ws.cell(row=r, column=3, value=date.fromisoformat(due))
            except ValueError:
                ws.cell(row=r, column=3, value=due)
        ws.cell(row=r, column=4, value=row.get("span", ""))
        ws.cell(row=r, column=5, value=row.get("note", ""))
        ws.cell(row=r, column=6, value="✓" if row.get("done") == "1" else "")
        ws.cell(row=r, column=7, value=(row.get("created_at") or "")[:10])

    for r in range(first, last + 1):
        for col in range(1, len(HEADERS) + 1):
            c = ws.cell(row=r, column=col)
            c.border = box
            c.font = Font(name=F, size=11, color=INK)
            c.alignment = Alignment(vertical="center",
                                    horizontal="left" if col in (2, 5) else "center",
                                    wrap_text=(col in (2, 5)),
                                    indent=1 if col in (2, 5) else 0)
            if (r - first) % 2 == 1:
                c.fill = PatternFill("solid", fgColor=BAND)
        ws.cell(row=r, column=3).number_format = "yyyy/m/d"
        ws.row_dimensions[r].height = 28

    dv_kind = DataValidation(type="list", formula1='"院長から,受付から,定例"',
                             allow_blank=True, showDropDown=False)
    dv_done = DataValidation(type="list", formula1='"✓"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_kind)
    ws.add_data_validation(dv_done)
    dv_kind.add(f"A{first}:A{last}")
    dv_done.add(f"F{first}:F{last}")

    rng = f"A{first}:G{last}"
    ws.conditional_formatting.add(rng, Rule(
        type="expression", formula=[f'$F{first}="✓"'],
        dxf=DifferentialStyle(font=Font(strike=True, color=FAINT)), stopIfTrue=True))
    ws.conditional_formatting.add(rng, Rule(
        type="expression", formula=[f'AND($C{first}<>"",$C{first}<TODAY())'],
        dxf=DifferentialStyle(font=Font(color=RED, bold=True),
                              fill=PatternFill(start_color=REDBG, end_color=REDBG,
                                               fill_type="solid")),
        stopIfTrue=False))
    ws.conditional_formatting.add(rng, Rule(
        type="expression", formula=[f'$A{first}="院長から"'],
        dxf=DifferentialStyle(fill=PatternFill(start_color=AMBERBG, end_color=AMBERBG,
                                               fill_type="solid")),
        stopIfTrue=False))

    for col, width in zip("ABCDEFG", WIDTHS):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = f"A{first}"
    ws.auto_filter.ref = f"A{HDR}:G{last}"
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = f"{HDR}:{HDR}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_properties.tabColor = "0F5F55"


def build_summary(ws, groups):
    ws["A1"] = "受付　TO DO リスト"
    ws["A1"].font = Font(name=F, size=16, bold=True, color=INK)
    ws["A2"] = "Chatworkのマイチャットに送った内容が、月ごとのシートに自動で貯まります。"
    ws["A2"].font = Font(name=F, size=10, color=MUTED)

    for i, head in enumerate(["月", "件数", "完了", "未完了"], start=1):
        c = ws.cell(row=4, column=i, value=head)
        c.font = Font(name=F, size=10.5, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=TEAL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = box

    for i, (title, rows) in enumerate(groups):
        r = 5 + i
        done = sum(1 for x in rows if x.get("done") == "1")
        for col, value in enumerate([title, len(rows), done, len(rows) - done], start=1):
            c = ws.cell(row=r, column=col, value=value)
            c.border = box
            c.font = Font(name=F, size=11, color=INK)
            c.alignment = Alignment(horizontal="left" if col == 1 else "center",
                                    vertical="center", indent=1 if col == 1 else 0)
            if i % 2 == 1:
                c.fill = PatternFill("solid", fgColor=BAND)

    if not groups:
        ws["A5"] = "まだ登録がありません"
        ws["A5"].font = Font(name=F, size=11, color=FAINT)

    for col, width in zip("ABCD", [18, 12, 12, 12]):
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1A2422"


def main():
    rows = todo_store.load()
    buckets = defaultdict(list)
    for row in rows:
        buckets[month_key(row)].append(row)

    def sort_key(name):
        year, month = name.replace("月", "").split("年")
        return (int(year), int(month))

    groups = [(name, buckets[name]) for name in sorted(buckets, key=sort_key)]

    wb = Workbook()
    build_summary(wb.active, groups)
    wb.active.title = "サマリー"
    for title, group in groups:
        build_sheet(wb.create_sheet(title), title, group)
    if groups:
        wb.active = wb.sheetnames.index(groups[-1][0])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[OK] {OUT.relative_to(ROOT)} を更新しました（{len(rows)}件）")


if __name__ == "__main__":
    main()
