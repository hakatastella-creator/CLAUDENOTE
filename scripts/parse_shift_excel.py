"""博多ステラ歯科のシフト表Excelをパースして中間CSVに変換する。

入力: カレンダー形式のExcel（月ごとにシート）
出力: date, name, slot, raw_note の中間CSV

slot は FULL / AM / PM / SPECIAL / CLOSED のいずれか。
SPECIAL は「17時まで」「15:30～」など個別時間指定の場合に使う。
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path

import openpyxl

# シフト表に登場する正式な氏名（姓のみ）。揺れがあれば右辺へ正規化する。
NAME_ALIASES = {
    "美景先生": "美景",
    "美景": "美景",
}
KNOWN_NAMES = {
    "永渕", "細井", "藤垣", "若松", "美景", "城戸",
    "柳田", "有田", "安武", "内倉",
    "百済", "阿比留", "片山",
}

# 曜日列 → 曜日番号（月=0）
WEEKDAY_COLS = {"B": 0, "D": 1, "F": 2, "H": 3, "J": 4, "L": 5, "N": 6}


@dataclass
class ShiftEntry:
    day: date
    name: str
    slot: str  # FULL / AM / PM / SPECIAL / CLOSED
    raw_note: str = ""


def normalize_name(token: str) -> str | None:
    """トークンから氏名（姓）を取り出して正規化する。"""
    # 全角空白を半角に
    token = token.replace("　", " ").strip()
    if not token:
        return None
    # 「先生」を取り除く
    token = re.sub(r"先生", "", token)
    # AM/PM や時間表記を取り除いた残りで姓を判定
    base = re.sub(r"(AM|PM|am|pm|\d+[:：]?\d*時?まで|\d+[:：]\d+～?)", "", token).strip("（）()・,， ")
    base = base.strip()
    if not base:
        return None
    base = NAME_ALIASES.get(base, base)
    if base in KNOWN_NAMES:
        return base
    return None


def classify_slot(token: str) -> tuple[str, str]:
    """トークンから (slot, note) を返す。"""
    note_match = re.search(r"[（(].*?[）)]", token)
    note = note_match.group(0) if note_match else ""

    # 個別時間指定（例: 17時まで, 15:30～）
    if re.search(r"\d+[:：]?\d*時?まで|\d+[:：]\d+\s*～", token):
        return "SPECIAL", note or token.strip()
    if "AM" in token or "am" in token:
        return "AM", note
    if "PM" in token or "pm" in token:
        return "PM", note
    return "FULL", note


def parse_cell(text: str, day: date) -> list[ShiftEntry]:
    """1セル分のテキストから ShiftEntry のリストを返す。"""
    entries: list[ShiftEntry] = []
    text = text.strip()
    if not text:
        return entries
    if "休診" in text:
        # 休診そのものはシフト不要だが、明示的に CLOSED 行を作っておく
        return [ShiftEntry(day=day, name="(休診)", slot="CLOSED")]

    # 改行・全角空白・半角空白でグループに分割した上で、各グループを「・」で分割
    # （グループの境界は意味的に「医師/衛生士/受付」だが、CSV出力では区別不要）
    chunks = re.split(r"[\n　 ]+", text)
    for chunk in chunks:
        if not chunk:
            continue
        for token in chunk.split("・"):
            token = token.strip()
            if not token:
                continue
            name = normalize_name(token)
            if name is None:
                continue
            slot, note = classify_slot(token)
            entries.append(ShiftEntry(day=day, name=name, slot=slot, raw_note=note))
    return entries


def parse_month_sheet(ws, year: int, target_month: int) -> list[ShiftEntry]:
    """1シート（1か月分）をパースする。

    レイアウト:
      日付行: B/D/F/H/J/L/N 列、行番号 5,7,9,11,13(,15)
      シフト行: 日付行の +1 行（B6, B8, ...）
    """
    entries: list[ShiftEntry] = []
    for date_row in range(5, ws.max_row + 1, 2):
        shift_row = date_row + 1
        for col, _wday in WEEKDAY_COLS.items():
            date_cell = ws[f"{col}{date_row}"]
            shift_cell = ws[f"{col}{shift_row}"] if shift_row <= ws.max_row else None
            day_num = date_cell.value
            if not isinstance(day_num, int):
                continue
            # 前月・翌月のはみ出し日は対象月のみに絞る
            # シート内で日番号が小さい→大きい→小さいと推移するので、
            # シフトが書かれているかどうかではなく日番号の月跨ぎを判定する。
            month_for_day = _infer_month(date_cell, year, target_month)
            if month_for_day != target_month:
                continue
            if shift_cell is None or shift_cell.value is None:
                continue
            day = date(year, target_month, day_num)
            entries.extend(parse_cell(str(shift_cell.value), day))
    return entries


def _infer_month(date_cell, year: int, target_month: int) -> int:
    """セルの位置と日番号から、その日が target_month に属するか判定する。"""
    day_num = date_cell.value
    if not isinstance(day_num, int):
        return -1
    row = date_cell.row
    # 1行目（row 5）かつ日番号が大きい(>=15)なら前月
    if row == 5 and day_num >= 15:
        return target_month - 1 if target_month > 1 else 12
    # 最終週付近かつ日番号が小さい(<=14)なら翌月
    # シート内の最大日付行を見て判定する
    ws = date_cell.parent
    last_date_row = max(
        (c.row for c in ws[date_cell.column_letter] if isinstance(c.value, int)),
        default=row,
    )
    if row == last_date_row and day_num <= 14 and row > 5:
        return target_month + 1 if target_month < 12 else 1
    return target_month


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="入力Excelファイル")
    ap.add_argument("--year", type=int, required=True)
    ap.add_argument("--month", type=int, required=True)
    ap.add_argument("--sheet", help="シート名（省略時は '<month>月'）")
    ap.add_argument("--output", required=True, help="出力CSV")
    args = ap.parse_args()

    sheet_name = args.sheet or f"{args.month}月"
    wb = openpyxl.load_workbook(args.input, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"シートが見つかりません: {sheet_name} (候補: {wb.sheetnames})")
    ws = wb[sheet_name]

    entries = parse_month_sheet(ws, args.year, args.month)
    entries.sort(key=lambda e: (e.day, e.name, e.slot))

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["date", "name", "slot", "raw_note"])
        for e in entries:
            w.writerow([e.day.isoformat(), e.name, e.slot, e.raw_note])

    # サマリも標準出力に
    print(f"出力: {out_path} ({len(entries)}件)")
    by_name: dict[str, int] = {}
    for e in entries:
        by_name[e.name] = by_name.get(e.name, 0) + 1
    for name, n in sorted(by_name.items(), key=lambda x: -x[1]):
        print(f"  {name}: {n}件")


if __name__ == "__main__":
    main()
